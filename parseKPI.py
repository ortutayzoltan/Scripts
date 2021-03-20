#!/usr/bin/python
import csv
import re , os
from openpyxl import Workbook
from optparse import OptionParser


types = {
'sip':['Timestamp', 'PmpId', 'NetId', 'NetType', 'IPversion', 'RegUsers', 'InitRegSuccRatio', 'InitRegTime', 'InitRegRate', 'ReRegRate', 'Sessions', 'IncSessionRate', 'OutSessionRate', 'IncSessSuccRatio', 'OutSessSuccRatio', 'IncSessSessionSetupTime', 'OutSessSessionSetupTime', 'SubscSessions', 'IncSubscRate', 'OutSubscRate', 'IncSubscSuccRatio', 'OutSubscSuccRatio', 'ThrottRatio '],
'bgf':['Timestamp', 'PmpId', 'BgfName', 'BgfSuccRatio', 'BgfContexts                                                                                                                                                                                                                                                                     '],
'dia1':['Timestamp', 'PmpId', 'DiameterRole', 'Realm', 'Instance', 'Role', 'SuccRatio                                                                                                                                                                                                                                                           '],
'dia2':['Timestamp', 'PmpId', 'DiameterRole', 'Realm', 'PrimaryInstance', 'SecondaryInstance', 'RxSuccNpliRatio                                                                                                                                                                                                                                 '],
'pmp':['Timestamp', 'PmpId', 'CpuLoadCh', 'CpuLoadSb', 'MemoryLoadCh', 'MemoryLoadSb', 'CpRegUsers', 'CpSessions                                                                                                                                                                                                                                  ']
}

parser = OptionParser()
parser.add_option("-f", "--file", dest="filename", help="filename", metavar="filename")
(options, args) = parser.parse_args()

if (options.filename is not None):
	filename = options.filename
else:
	filename = 'kpi.csv'

#Timestamp', 'PmpId', 'CpuLoadCh', 'CpuLoadSb', 'MemoryLoadCh', 'MemoryLoadSb', 'CpRegUsers', 'CpSessions

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "CpuLoadCh"
ws2 = wb.create_sheet(title="CpuLoadSb")
ws3 = wb.create_sheet(title="MemoryLoadCh")
ws4 = wb.create_sheet(title="MemoryLoadSb")
ws5 = wb.create_sheet(title="CpRegUsers")
ws6 = wb.create_sheet(title="CpSessions")
ws = None
CpuLoadCh    = []
CpuLoadSb    = []
MemoryLoadCh = []
MemoryLoadSb = []
CpRegUsers   = []
CpSessions   = []
with open(filename) as csvfile:
	rows = csv.reader(csvfile)
	for row in rows:
		if len(row) != 8 :
			continue
			
		if row[1] == 'pmp_1':
			CpuLoadCh    = [row[0]]
			CpuLoadSb    = [row[0]]
			MemoryLoadCh = [row[0]]
			MemoryLoadSb = [row[0]]
			CpRegUsers   = [row[0]]
			CpSessions   = [row[0]]
		CpuLoadCh.append(row[2])
		CpuLoadSb.append(row[3])
		MemoryLoadCh.append(row[4])
		MemoryLoadSb.append(row[5])
		CpRegUsers.append(row[6])
		CpSessions.append(row[7])
		if row[1] == 'pmp_10':
			ws1.append(CpuLoadCh    )
			ws2.append(CpuLoadSb    )
			ws3.append(MemoryLoadCh )
			ws4.append(MemoryLoadSb )
			ws5.append(CpRegUsers   )
			ws6.append(CpSessions   )


wb.save(filename = 'KPI.xlsx')