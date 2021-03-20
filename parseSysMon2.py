#!/usr/bin/python3
import csv
import re , os
from openpyxl import Workbook
from optparse import OptionParser

parser = OptionParser()
parser.add_option("-f", "--file", dest="filename", help="filename used for processing, if omitted sysMon2small.1 is used. If omitted all files will be processed from the current folder", metavar="filename")
(options, args) = parser.parse_args()

if (options.filename is not None):
	filenames = [options.filename]
else:
	filenames = [f for f in os.listdir('./') if re.match(r'^sysMon2small.\d$',f)]

records = [ 'date', 'time', 'cp_stat_cpu', 'cp_stat_all_cpu', 'cp_stat_tot_free', 'cpu', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', 'avg_erlang_cpu', 'total', 'processes', 'system', 'atom', 'binary', 'code', 'ets', 'conntrack_count', 'run_queue', 'reg_procs', 'reg_bindings', 'reg_errs', 'sip_transport', 'b2bDialogTable_it', 'b2bDialogTable_ib', 'b2bDialogTable_ot', 'b2bDialogTable_ob', 'sip_ack', 'plc_fg', 'plc_bg', 'sysIPsec']
wb = Workbook()
first_sheet = wb.get_sheet_by_name('Sheet')

first_sheet.append(['cp_stat'])
first_sheet.append(['cp_stat_cpu, cp_stat_all_cpu, cp_stat_tot_free', 'CPU figures from sysPerf. These are the ones displayed in PERF.'])
first_sheet.append([''])
first_sheet.append(['cpu'])
first_sheet.append(['cpu','cpu utilization. Based on erlang:statistics wallclock and runtime'])
first_sheet.append([''])
first_sheet.append(['cores'])
first_sheet.append(['1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16','CPU load per core'])
first_sheet.append([''])
first_sheet.append(['avg_erlang_cpu'])
first_sheet.append(['avg_erlang_cpu' , 'shows HT avg_erlang_cpu utilization. Based on erlang:statistics(scheduler_wall_time). request method names, emergency priorities are grouped together. This counter are calculated the same way as counter sbgApplCPUUsage and  show the average erlang cpu load during the last sysMon2 log interval'])
first_sheet.append([''])
first_sheet.append(['Memory usage'])
first_sheet.append(['total, processes, system, atom, binary, code, ets' , 'Prints the output of erlang:memory/0, minus processes_used and atoms_used'])
first_sheet.append([''])
first_sheet.append(['Total number of connections'])
first_sheet.append(['conntrack_count', 'Total number of connections from  /proc/sys/net/netfilter/nf_conntrack_count'])
first_sheet.append([''])
first_sheet.append(['run_que'])
first_sheet.append(['run_que' , 'erlang VM run_queue. PLC uses this for load regulating (among other things). A high number indicates high actual load. This is the number of processes waiting to execute, VM-wide'])
first_sheet.append([''])
first_sheet.append(['reg counters'])
first_sheet.append(['reg_procs, reg_bindings, reg_errs', 'counters from REG. Number of of REG CCPC processes (one per IMPU), Number of bindings (differ from procs if more than one contact per IMPU), Number of errors in REG'])
first_sheet.append([''])
first_sheet.append(['Tables'])
first_sheet.append(['sip_transport, b2bDialogTable_it, b2bDialogTable_ib, b2bDialogTable_ot, b2bDialogTable_ob, sip_ack, plc_fg, plc_bg', 'Table sizes of these tables'])
first_sheet.append([''])
first_sheet.append(['sysIPsec'])
first_sheet.append(['sysIPsec' , 'Display the message queue length of processes, sysIPsec'])
first_sheet.append([])

sheetnames = ['cp_stat','cpu','cores','avg_erlang_cpu','Memory usage','Total number of connections','run_que','reg counters','Tables','sysIPsec']
sheets = [wb.create_sheet(title=name) for name  in sheetnames]
date = ''

def convert_to_float(string):
	try:
		retval = float(string)
	except:
		retval = string
	return retval
# Addig headers the ugly way
sheets[0].append(['Timestamp', 'cp_stat_cpu', 'cp_stat_all_cpu', 'cp_stat_tot_free'])
sheets[1].append(['Timestamp', 'CPU'])
sheets[2].append(['Timestamp', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16'])
sheets[3].append(['Timestamp', 'avg_erlang_cpu'])
sheets[4].append(['Timestamp', 'total', 'processes', 'system', 'atom', 'binary', 'code', 'ets'])
sheets[5].append(['Timestamp', 'conntrack_count'])
sheets[6].append(['Timestamp', 'run_que'])
sheets[7].append(['Timestamp', 'reg_procs', 'reg_bindings', 'reg_errs'])
sheets[8].append(['Timestamp', 'sip_transport', 'b2bDialogTable_it', 'b2bDialogTable_ib', 'b2bDialogTable_ot', 'b2bDialogTable_ob', 'sip_ack', 'plc_fg', 'plc_bg'])
sheets[9].append(['Timestamp', 'sysIPsec'])

for filename in filenames:
	print(filename)
	with open(filename) as csvfile:
		rows = csv.reader(csvfile)
		for row in rows:
			if row == records or len(row)<4:
				continue
			else:
				if row[0]!='':
					date = row[0]
				if date == '':
					continue
				timeStamp = date + 'T' + row[1]
				sheets[0].append([timeStamp,convert_to_float(row[2]),convert_to_float(row[3]),convert_to_float(row[4])])
				sheets[1].append([timeStamp,convert_to_float(row[5])])
				sheets[2].append([timeStamp,convert_to_float(row[6]),convert_to_float(row[7]),convert_to_float(row[8]),convert_to_float(row[9]),convert_to_float(row[10]),convert_to_float(row[11]),convert_to_float(row[12]),convert_to_float(row[13]),convert_to_float(row[14]),convert_to_float(row[15]),convert_to_float(row[16]),convert_to_float(row[17]),convert_to_float(row[18]),convert_to_float(row[19]),convert_to_float(row[20]),convert_to_float(row[21])])
				sheets[3].append([timeStamp,convert_to_float(row[22])])
				sheets[4].append([timeStamp, convert_to_float(row[23]) , convert_to_float(row[24]), convert_to_float(row[25]), convert_to_float(row[26]), convert_to_float(row[27]), convert_to_float(row[28]), convert_to_float(row[29])])
				sheets[5].append([timeStamp, convert_to_float(row[30])] )
				sheets[6].append([timeStamp, convert_to_float(row[31])] )
				sheets[7].append([timeStamp, convert_to_float(row[32]) , convert_to_float(row[33]), convert_to_float(row[34])])
				sheets[8].append([timeStamp, convert_to_float(row[35]) , convert_to_float(row[36]), convert_to_float(row[37]), convert_to_float(row[38]), convert_to_float(row[39]), convert_to_float(row[40]), convert_to_float(row[41]), convert_to_float(row[42])])
				sheets[9].append([timeStamp, convert_to_float(row[43])] )
			

wb.save(filename = 'sysMon2small.xlsx')

			 

		