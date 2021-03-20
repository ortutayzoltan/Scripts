#!/usr/bin/python
import csv
import re , os
from openpyxl import Workbook
from optparse import OptionParser


types = {
'sip':['Timestamp', 'PmpId', 'NetId', 'NetType', 'IPversion', 'RegUsers', 'InitRegSuccRatio', 'InitRegTime', 'InitRegRate', 'ReRegRate', 'Sessions', 'IncSessionRate', 'OutSessionRate', 'IncSessSuccRatio', 'OutSessSuccRatio', 'IncSessSessionSetupTime', 'OutSessSessionSetupTime', 'SubscSessions', 'IncSubscRate', 'OutSubscRate', 'IncSubscSuccRatio', 'OutSubscSuccRatio', 'ThrottRatio '],
'bgf':['Timestamp', 'PmpId', 'BgfName', 'BgfSuccRatio', 'BgfContexts                                                                                                                                                                                                                                                                     '],
'dia1':['Timestamp', 'PmpId', 'DiameterRole', 'Realm', 'Instance', 'Role', 'SuccRatio                                                                                                                                                                                                                                                           '],
'dia2':['Timestamp', 'PmpId', 'DiameterRole', 'Realm', 'PrimaryInstance', 'SecondaryInstance', 'RxSuccNpliRatio                                                                                                                                                                                                                                 '],
'pmp':['Timestamp', 'PmpId', 'CpuLoadCh', 'CpuLoadSb', 'MemoryLoadCh', 'MemoryLoadSb', 'CpRegUsers', 'CpSessions                                                                                                                                                                                                                                  ']
}

parser = OptionParser()
parser.add_option("-f", "--file", dest="filename", help="filename", metavar="filename")
(options, args) = parser.parse_args()

if (options.filename is not None):
	filename = options.filename
else:
	filename = 'kpi.csv'

def convert_to_float(string):
	try:
		retval = float(string)
	except:
		retval = string
	return retval

#Timestamp', 'PmpId', 'CpuLoadCh', 'CpuLoadSb', 'MemoryLoadCh', 'MemoryLoadSb', 'CpRegUsers', 'CpSessions

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "CpuLoadCh"
ws2 = wb.create_sheet(title="CpuLoadSb")
ws3 = wb.create_sheet(title="MemoryLoadCh")
ws4 = wb.create_sheet(title="MemoryLoadSb")
ws5 = wb.create_sheet(title="CpRegUsers")
ws6 = wb.create_sheet(title="CpSessions")
ws = None

ws1.append(['timeStamp' , 'pmp' , 'CpuLoadCh'])
ws2.append(['timeStamp' , 'pmp' , 'CpuLoadSb'])
ws3.append(['timeStamp' , 'pmp' , 'MemoryLoadCh'])
ws4.append(['timeStamp' , 'pmp' , 'MemoryLoadSb'])
ws5.append(['timeStamp' , 'pmp' , 'CpRegUsers'])
ws6.append(['timeStamp' , 'pmp' , 'CpSessions'])

ws1.auto_filter.ref = "A1:C1048576"
ws2.auto_filter.ref = "A1:C1048576"
ws3.auto_filter.ref = "A1:C1048576"
ws4.auto_filter.ref = "A1:C1048576"
ws5.auto_filter.ref = "A1:C1048576"
ws6.auto_filter.ref = "A1:C1048576"

with open(filename) as csvfile:
	rows = csv.reader(csvfile)
	for row in rows:
		if len(row) != 8 or row[0] == 'Timestamp':
			continue
		print(row)
		timeStamp =    row[0]
		pmp =          row[1]
		CpuLoadCh =    row[2]
		CpuLoadSb =    row[3]
		MemoryLoadCh = row[4]
		MemoryLoadSb = row[5] 
		CpRegUsers =   row[6]
		CpSessions =   row[7]
		ws1.append([timeStamp , pmp , convert_to_float(CpuLoadCh    )])
		ws2.append([timeStamp , pmp , convert_to_float(CpuLoadSb    )])
		ws3.append([timeStamp , pmp , convert_to_float(MemoryLoadCh )])
		ws4.append([timeStamp , pmp , convert_to_float(MemoryLoadSb )])
		ws5.append([timeStamp , pmp , convert_to_float(CpRegUsers   )])
		ws6.append([timeStamp , pmp , convert_to_float(CpSessions   )])


wb.save(filename = 'KPI.xlsx')